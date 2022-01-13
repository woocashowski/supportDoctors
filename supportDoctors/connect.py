import math
from rsa import encrypt
from rsa import decrypt
from rsa import generateKeys

import openpyxl
from pathlib import Path

from sklearn.neighbors import KNeighborsClassifier

import numpy as np


def read_patient(id_p):
    xlsx_file = Path('patients.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    print("  ")
    id = sheet.cell(row=id_p, column=1).value
    password = sheet.cell(row=id_p, column=2).value
    k1 = sheet.cell(row=id_p, column=3).value
    k2 = sheet.cell(row=id_p, column=4).value
    k3 = sheet.cell(row=id_p, column=5).value
    k4 = sheet.cell(row=id_p, column=6).value
    k5 = sheet.cell(row=id_p, column=7).value
    k6 = sheet.cell(row=id_p, column=8).value
    k7 = sheet.cell(row=id_p, column=9).value
    k8 = sheet.cell(row=id_p, column=10).value
    outcome = sheet.cell(row=id_p, column=11).value
    prediction = sheet.cell(row=id_p, column=12).value

    patient_values = Patient(id, password, k1, k2, k3, k4,
                             k5, k6, k7, k8, outcome, prediction)

    return patient_values


def add_patient(patient_values):
    xlsx_file = Path('patients.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    # ID:
    last_empty_row = len(list(sheet.rows)) + 1
    # set ID
    patient_values.id = last_empty_row
    print("row: ", last_empty_row)
    # Add to database
    sheet.cell(row=last_empty_row, column=1).value = last_empty_row
    sheet.cell(row=last_empty_row, column=2).value = patient_values.password
    sheet.cell(row=last_empty_row, column=3).value = patient_values.k1
    sheet.cell(row=last_empty_row, column=4).value = patient_values.k2
    sheet.cell(row=last_empty_row, column=5).value = patient_values.k3
    sheet.cell(row=last_empty_row, column=6).value = patient_values.k4
    sheet.cell(row=last_empty_row, column=7).value = patient_values.k5
    sheet.cell(row=last_empty_row, column=8).value = patient_values.k6
    sheet.cell(row=last_empty_row, column=9).value = patient_values.k7
    sheet.cell(row=last_empty_row, column=10).value = patient_values.k8
    sheet.cell(row=last_empty_row, column=11).value = patient_values.outcome
    sheet.cell(row=last_empty_row, column=12).value = patient_values.prediction

    print("value: ", sheet.cell(row=last_empty_row, column=1).value)
    wb_obj.save(xlsx_file)


def edit_patient(patient_values):
    xlsx_file = Path('patients.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    # ID:
    patient_id = patient_values.id
    # Edit database
    sheet.cell(row=patient_id, column=2).value = patient_values.password
    sheet.cell(row=patient_id, column=3).value = patient_values.k1
    sheet.cell(row=patient_id, column=4).value = patient_values.k2
    sheet.cell(row=patient_id, column=5).value = patient_values.k3
    sheet.cell(row=patient_id, column=6).value = patient_values.k4
    sheet.cell(row=patient_id, column=7).value = patient_values.k5
    sheet.cell(row=patient_id, column=8).value = patient_values.k6
    sheet.cell(row=patient_id, column=9).value = patient_values.k7
    sheet.cell(row=patient_id, column=10).value = patient_values.k8
    sheet.cell(row=patient_id, column=11).value = patient_values.outcome
    sheet.cell(row=patient_id, column=12).value = patient_values.prediction
    # Save
    wb_obj.save(xlsx_file)


class Patient:
    def __init__(self, id, password, k1, k2, k3, k4,
                 k5, k6, k7, k8, outcome, prediction):
        self.id = id
        self.password = password
        self.k1 = k1
        self.k2 = k2
        self.k3 = k3
        self.k4 = k4
        self.k5 = k5
        self.k6 = k6
        self.k7 = k7
        self.k8 = k8
        self.outcome = outcome
        self.prediction = prediction


patient = Patient(0, 0, 0, 0, 0, 0,
                  0, 0, 0, 0, 0, 0)


# print(read_patient(769).id)
# patient = read_patient(1)
# print(read_patient(769).id)
# print(read_patient(770).id)
# print(read_patient(770).id)


# ==========================================================================
# kNN

def get_neighbor(id_pred):
    xlsx_file = Path('patients.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    X = []
    y = []
    last_empty_row = len(list(sheet.rows))
    last_empty_row = 769
    for i in range(2, last_empty_row - 1):
        try:
            X.append([
                float(sheet.cell(row=i, column=3).value),
                float(sheet.cell(row=i, column=4).value),
                float(sheet.cell(row=i, column=5).value),
                float(sheet.cell(row=i, column=6).value),
                float(sheet.cell(row=i, column=7).value),
                float(sheet.cell(row=i, column=8).value),
                float(sheet.cell(row=i, column=9).value),
                float(sheet.cell(row=i, column=10).value)
            ])
            y.append(sheet.cell(row=i, column=11).value)
            print(i)
        finally:
            pass

    neigh = KNeighborsClassifier(n_neighbors=3)
    neigh.fit(X, y)

    my_patient = read_patient(id_pred)
    new_patient = [
        my_patient.k1,
        my_patient.k2,
        my_patient.k3,
        my_patient.k4,
        my_patient.k5,
        my_patient.k6,
        my_patient.k7,
        my_patient.k8,
    ]
    A = neigh.predict([new_patient])
    print("A: ", A)
    return A


# get_neighbor(10)


# ================================================================================
# doctors
def read_keys():
    lines = []
    with open('pass.txt') as f:
        lines = f.readlines()
    f.close()
    return lines


def add_doctor(login, password):
    xlsx_file = Path('doctors.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    canAdd = True
    print(login, password)
    # ID:
    last_empty_row = len(list(sheet.rows)) + 1
    # check login in DB
    for i in range(2, last_empty_row):
        if sheet.cell(row=i, column=2).value == login:
            canAdd = False
            print(sheet.cell(row=i, column=2).value, " ", i)
            print(login)

    # Add to database
    if (canAdd):
        e = int(read_keys()[0][0:-1])
        d = int(read_keys()[1][0:-1])
        N = int(read_keys()[2][0:-1])

        sheet.cell(row=last_empty_row, column=1).value = last_empty_row
        sheet.cell(row=last_empty_row, column=2).value = login
        sheet.cell(row=last_empty_row, column=3).value = encrypt(e, N, password)

    print("value: ", sheet.cell(row=last_empty_row, column=1).value)
    wb_obj.save(xlsx_file)
    return canAdd


def login_doctor(login, password):
    xlsx_file = Path('doctors.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    can_login = False
    print(login, password)
    # ID:
    last_empty_row = len(list(sheet.rows)) + 1
    # check login in DB
    e = int(read_keys()[0][0:-1])
    d = int(read_keys()[1][0:-1])
    N = int(read_keys()[2][0:-1])
    password = encrypt(e, N, password)
    for i in range(2, last_empty_row):
        print(i, " ", sheet.cell(row=i, column=2).value, " ", login, " pss ", sheet.cell(row=i, column=3).value,
              " pss2 ", password)
        if sheet.cell(row=i, column=2).value == login and sheet.cell(row=i, column=3).value == password:
            can_login = True
            print(sheet.cell(row=i, column=2).value, " ", i)
    wb_obj.save(xlsx_file)
    return can_login


add_doctor("admin", "admin")
login_doctor("adam", "XD")


def set_keys():
    keysize = 32
    e, d, N = generateKeys(keysize)
    print(e)
    print(d)
    print(N)
